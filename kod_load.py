# -*- coding: utf-8 -*-
# Сделал предварительный backup изменяемой таблицы

from mysql.connector import MySQLConnection, Error
import openpyxl
import sys
import time
import csv
from datetime import datetime
from lib import l, read_config, format_police_code

IN_SNILS = ['СНИЛС', 'СтраховойНомер', 'Страховой_номер', 'Страховой Номер', 'Номер СНИЛС']
IN_KOD = ['Код_подразделения', 'Код подразделения', 'Код', 'Код_подразделения_выдавшего_документ']

dbconfig = read_config(filename='kod_load.ini', section='mysql')
dbconn = MySQLConnection(**dbconfig)  # Открываем БД из конфиг-файла

fields = []
workbooks =  []
sheets = []
for i, xlsx_file in enumerate(sys.argv):                              # Загружаем все xlsx файлы
    if i == 0:
        continue
    workbooks.append(openpyxl.load_workbook(filename=xlsx_file, read_only=True))
    sheets.append(workbooks[i-1][workbooks[i-1].sheetnames[0]])

sheets_keys = []
total_rows = 0
for i, sheet in enumerate(sheets):                                    # Маркируем нужные столбцы
    total_rows += sheet.max_row
    keys = {}
    for j, row in enumerate(sheet.rows):
        if j > 0:
            break
        for k, cell in enumerate(row):                                # Проверяем, чтобы был СНИЛС и Код
            if cell.value in IN_SNILS:
                keys[IN_SNILS[0]] = k
            if cell.value in IN_KOD:
                keys[IN_KOD[0]] = k
        if len(keys) < 2:
            print('В файле "' + sys.argv[i+1] + '" отсутствует колонка со СНИЛС или кодом подразделеня')
            time.sleep(3)
            sys.exit()
    sheets_keys.append(keys)

print('\n'+ datetime.now().strftime("%H:%M:%S") +' Начинаем корректировку кодов подразделений \n')

# добавить ключи из p_check
write_rows = []
new_file = True
backup_tuple = tuple()
sql = 'SELECT * FROM clients WHERE clients.number IN ('
cl_csvs = []
for i, sheet in enumerate(sheets):                              # Загружаем все xlsx файлы по мере сохранения в БД
    for j, row in enumerate(sheet.rows):                              # Теперь строки
        if j == 0:
            continue
        if new_file:
            sql += '%s'
            new_file = False
        else:
            sql += ',%s'
        backup_tuple += (l(row[keys[IN_SNILS[0]]].value),)
        write_row = (format_police_code(row[keys[IN_KOD[0]]].value), l(row[keys[IN_SNILS[0]]].value))
        write_rows.append(write_row)
        if j % 10000 == 0:
            read_cursor = dbconn.cursor()
            read_cursor.execute(sql+');',backup_tuple)
            rows = read_cursor.fetchall()
            for row in rows:
                cl_csv = {}
                for k, name in enumerate(read_cursor.description):
                    cl_csv[name[0]] = row[k]
                cl_csvs.append(cl_csv)
            new_file = True
            sql = 'SELECT * FROM clients WHERE clients.number IN ('
            write_cursor = dbconn.cursor()
            write_cursor.executemany('UPDATE clients SET p_police_code = %s  WHERE clients.number = %s', write_rows)
            dbconn.commit()
            write_rows = []
            print(datetime.now().strftime("%H:%M:%S") + ' 10k из файла '+ xlsx_file +' загрузил')
    print('\n' + datetime.now().strftime("%H:%M:%S") + ' Файл '+ xlsx_file +' загружен полностью\n')

read_cursor = dbconn.cursor()
read_cursor.execute(sql + ');', backup_tuple)
rows = read_cursor.fetchall()
col_names = []
for name in read_cursor.description:
    col_names.append(name[0])
for row in rows:
    cl_csv = {}
    for k, name in enumerate(read_cursor.description):
        cl_csv[name[0]] = row[k]
    cl_csvs.append(cl_csv)
# new_file = True
# sql = 'SELECT * FROM clients WHERE clients.number IN ('

a = sys.argv[1]
if len(a.split('/')) > 1:
    path = '/'.join(a.split('/')[:len(a.split('/'))-1])+'/'              # только путь без имени файла
else:
    path = ''

with open(path + 'cl.csv', 'w', encoding='cp1251') as output_file:  # backup изменяемой таблицы
    dict_writer = csv.DictWriter(output_file, col_names, delimiter=';') #, quoting=csv.QUOTE_NONNUMERIC)
    dict_writer.writeheader()
    dict_writer.writerows(cl_csvs)
output_file.close()

write_cursor = dbconn.cursor()
write_cursor.executemany('UPDATE clients SET p_police_code = %s  WHERE clients.number = %s', write_rows)
dbconn.commit()

print('\n'+ datetime.now().strftime("%H:%M:%S") +' Корректировка окончена \n')